#!/usr/bin/env python
import pandas as pd
import os


def pd_formatting(df, dst_file, precision=2):
    
    # openpyxl - A Python library to read/write Excel 2010 xlsx/xlsm files
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment, Side, Border
    
    wb = Workbook()
    ws = wb.active
            
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    
    print("\nFormatting the excel file, adjust the column width and precision...")
    
    # column name cell setting
    for i in range(1,df.shape[1]+1):
        cell = ws.cell(row=1, column=i)
        cell.font = Font(name="Arial",size=10, bold=True,italic=False)
        cell.alignment = Alignment(horizontal="center",vertical="center")
        cell.border = Border(bottom=Side(border_style="thin",color='FF000000'))
                             
    #  content cell setting
    for i in range(2,df.shape[0]+2):
        for j in range(1,df.shape[1]+1):
            cell = ws.cell(row=i, column=j)
            cell.font = Font(name="Arial",size=10, bold=False,italic=False)
            cell.alignment = Alignment(horizontal="center",vertical="center")
            # number format
            if type(cell.value) == int:
                cell.number_format = '0'
            elif type(cell.value) == float:
                if precision == 2:
                    cell.number_format = '0.00'
                elif precision == 3:
                    cell.number_format = '0.000'
                elif precision == 4:
                    cell.number_format = '0.0000'
                else:
                    cell.number_format = '0.0'
            elif type(cell.value) == str:
                pass
    
    
    # row height setting
    # Todo: row height setting
    # row_height = 40
    # for i in range(1, ws.max_row+1):
    #     ws.row_dimensions[i].height = row_height
        
    # column width setting
    column_width = 15
    for i in range(1, 2):
        ws.column_dimensions[get_column_letter(i)].width = 25
    
    for i in range(2, ws.max_column+1):
        ws.column_dimensions[get_column_letter(i)].width = column_width
        
    wb.save(dst_file)



def pd_create_tabel(args):
    """
    Args:
        args: column name

    Returns:

    """
    data = {}
    if isinstance(args, list):
        for _ in args:
            data[str(_)] = []

    return data


def pd_write_table(data, return_dataframe=False, save_excel=True, excel_path=None, formatting=True, precision=2):

    if not isinstance(data, dict):
        raise TypeError("Input is expected to be dict, not {}".format(type(data)))
    
    df = pd.DataFrame()
    for k in data.keys():
        df = pd.concat([df, pd.DataFrame({k: data[k]})], axis=1)
    
    # https://pandas.pydata.org/docs/reference/api/pandas.DataFrame.to_excel.html
    if save_excel:
        if formatting:
            pd_formatting(df, excel_path, precision)
        else:
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False) 
            
        print("\nSave result to excel: {0} \n".format(excel_path))

    if return_dataframe:
        return df

def generate_dataframe(res, column_name, return_dataframe=False, save_excel=True, excel_path=None, excel_formatting=True, 
                       precision=2):
    """
    res: list include row data
    column_name: list include column name
    """

    if isinstance(res, list) and isinstance(column_name, list):
        
        assert len(res[0]) == len(column_name), "The length of res and column should be the same!"
        
        
        data = pd_create_tabel(column_name)
        for i in range(len(res)):
            for j in range(len(column_name)):
                data[column_name[j]].append(res[i][j])

        df = pd_write_table(data, return_dataframe=return_dataframe, save_excel=save_excel, 
                            excel_path=excel_path, precision=precision, formatting=excel_formatting)
    else:
        raise TypeError("Input is expected to be both list, now is {} and {}".format(type(res), type(column_name)))

    if return_dataframe:
        return df
    

if __name__ == '__main__':
    
    x = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
    y = ['a', 'b', 'c']
    df = generate_dataframe(x, y, excel_path='test.xlsx', save_excel=False,return_dataframe=True,)
    print(df)
