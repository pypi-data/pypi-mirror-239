import json
from tornado import gen, web

from snb_node.base.Interceptor import interceptor
from snb_node.tools.utils import maybe_future, url_path_join, url_escape
from snb_node.contents.filemanager import FileContentsManager
import io
import os
import csv
import time
from decimal import Decimal
from openpyxl import load_workbook
import pandas as pd
try:
    from jupyter_client.jsonutil import json_default
except ImportError:
    from jupyter_client.jsonutil import (
        date_default as json_default
    )

from snb_node.base.handlers import (
    IPythonHandler, APIHandler, path_regex,
)


def validate_model(model, expect_content):
    """
    Validate a model returned by a ContentsManager method.

    If expect_content is True, then we expect non-null entries for 'content'
    and 'format'.
    """
    required_keys = {
        "name",
        "file_path",
        "type",
        "created",
        "update_time",
        "content",
        "format",
    }
    missing = required_keys - set(model.keys())
    if missing:
        raise web.HTTPError(
            500,
            f"Missing Model Keys: {missing}",
        )

    maybe_none_keys = ['content', 'format']
    if expect_content:
        errors = [key for key in maybe_none_keys if model[key] is None]
        if errors:
            raise web.HTTPError(
                500,
                f"Keys unexpectedly None: {errors}",
            )
    else:
        errors = {
            key: model[key]
            for key in maybe_none_keys
            if model[key] is not None
        }
        if errors:
            raise web.HTTPError(
                500,
                f"Keys unexpectedly not None: {errors}",
            )


class ContentsHandler(APIHandler):

    def location_url(self, path):
        return url_path_join(
            self.base_url, 'api', 'contents', url_escape(path)
        )

    def _finish_model(self, model, location=True):
        """Finish a JSON request with a model, setting relevant headers, etc."""
        if location:
            location = self.location_url(model['path'])
            self.set_header('Location', location)
        self.set_header('Content-Type', 'application/json')
        self.finish(json.dumps(model, default=json_default))

    @web.authenticated
    @gen.coroutine
    @interceptor
    def get(self, ws_uid, path='') -> {"GRADE": ["BASIC", "PRO", "ENT"], "ROLE": ["ADMIN", "EDITOR"]}:
        """Return a model for a file or directory.

        A directory model contains a list of models (without content)
        of the files and directories it contains.
        """
        path = path or ''
        cm = FileContentsManager()
        type = None
        """self.get_query_argument('type', default=None)
        if type not in {None, 'directory', 'file', 'notebook'}:
            raise web.HTTPError(400, f'Type {type!r} is invalid')

        format = self.get_query_argument('format', default=None)
        if format not in {None, 'text', 'base64'}:
            raise web.HTTPError(400, f'Format {format!r} is invalid')
        content = self.get_query_argument('content', default='1')
        if content not in {'0', '1'}:
            raise web.HTTPError(400, f'Content {content!r} is invalid')"""
        content = -1
        #file_name = self.get_argument('file_name', '')
        model = yield maybe_future(cm.get(
            path=path, type=type, format=format, content=content
        ))

        validate_model(model, expect_content=content)

        ret_obj={
            "code": 200,
            "data": model,
            "msg": "文件List成功！"
        }
        self._finish_model(ret_obj, location=False)

    @web.authenticated
    @gen.coroutine
    def patch(self, path=''):
        """PATCH renames a file or directory without re-uploading content."""
        cm = self.contents_manager
        model = self.get_json_body()
        old_path = model.get('path')
        if old_path and (cm.is_hidden(path) or cm.is_hidden(old_path))  and not cm.allow_hidden:
            raise web.HTTPError(400, f'Cannot rename file or directory {path!r}')
        if model is None:
            raise web.HTTPError(400, 'JSON body missing')
        model = yield maybe_future(cm.update(model, path))
        validate_model(model, expect_content=False)
        self._finish_model(model)

    @gen.coroutine
    def _copy(self, copy_from, copy_to=None):
        """Copy a file, optionally specifying a target directory."""
        self.log.info(f"Copying {copy_from} to {copy_to or ''}")
        model = yield maybe_future(self.contents_manager.copy(copy_from, copy_to))
        self.set_status(201)
        validate_model(model, expect_content=False)
        self._finish_model(model)

    @gen.coroutine
    def _upload(self, model, path):
        """Handle upload of a new file to path"""
        self.log.info("Uploading file to %s", path)
        model = yield maybe_future(self.contents_manager.new(model, path))
        self.set_status(201)
        validate_model(model, expect_content=False)
        self._finish_model(model)

    @gen.coroutine
    def _new_untitled(self, path, type='', ext=''):
        """Create a new, empty untitled entity"""
        self.log.info("Creating new %s in %s", type or 'file', path)
        model = yield maybe_future(self.contents_manager.new_untitled(path=path, type=type, ext=ext))
        self.set_status(201)
        validate_model(model, expect_content=False)
        self._finish_model(model)

    @gen.coroutine
    def _save(self, model, path):
        """Save an existing file."""
        chunk = model.get("chunk", None)
        if not chunk or chunk == -1:  # Avoid tedious log information
            self.log.info("Saving file at %s", path)
        model = yield maybe_future(self.contents_manager.save(model, path))
        validate_model(model, expect_content=False)
        self._finish_model(model)

    @web.authenticated
    @gen.coroutine
    def post(self, ws_uid, path=''):
        try:
            # 提取表单中‘name’为‘files’的文件元数据
            files = self.request.files
            if path.startswith("/home"):
                file_path = path
            else:
                file_path='/home/'+path
            ret_str=''
            for file in files:
                meta = files[file][0]
                filename = meta["filename"]
                if os.path.exists(file_path+"/"+filename):
                    ret_str=ret_str+"\n"+filename+" 存在上传失败！"
                    ret_obj = {
                        "code": 400,
                        "msg": ret_str
                    }
                else:
                    with open(file_path+"/"+filename,'wb+') as fp:
                        fp.write(meta["body"])
                    ret_str = ret_str + "\n" + filename + " 上传成功！"
                    ret_obj = {
                        "code": 200,
                        "msg": ret_str
                    }
            self._finish_model(ret_obj, location=False)
        except Exception as e:
            ret_obj = {
                "code": 400,
                "msg": "文件上传失败："+str(e)
            }
            self._finish_model(ret_obj, location=False)

        """Create a new file in the specified path.

        POST creates new files. The server always decides on the name.

        POST /api/contents/path
          New untitled, empty file or directory.
        POST /api/contents/path
          with body {"copy_from" : "/path/to/OtherNotebook.ipynb"}
          New copy of OtherNotebook in path
        
        cm = FileContentsManager()

        file_exists = yield maybe_future(cm.file_exists(path))
        if file_exists:
            raise web.HTTPError(400, "Cannot POST to files, use PUT instead.")

        dir_exists = yield maybe_future(cm.dir_exists(path))
        if not dir_exists:
            raise web.HTTPError(404, f"No such directory: {path}")

        model = self.get_json_body()
        copy_from = model.get('copy_from')
        if copy_from and (cm.is_hidden(path) or cm.is_hidden(copy_from))  and not cm.allow_hidden:
            raise web.HTTPError(400, f'Cannot copy file or directory {path!r}')

        if model is not None:
            copy_from = model.get('copy_from')
            ext = model.get('ext', '')
            type = model.get('type', '')
            if copy_from:
                yield self._copy(copy_from, path)
            else:
                yield self._new_untitled(path, type=type, ext=ext)
        else:
            yield self._new_untitled(path)"""

    @web.authenticated
    @gen.coroutine
    def put(self, ws_uid, path=''):
        """Saves the file in the location specified by name and path.

        PUT is very similar to POST, but the requester specifies the name,
        whereas with POST, the server picks the name.

        PUT /api/contents/path/Name.ipynb
          Save notebook at ``path/Name.ipynb``. Notebook structure is specified
          in `content` key of JSON request body. If content is not specified,
          create a new empty notebook.
        """
        model = self.get_json_body()
        cm = FileContentsManager()
        if model:
            if model.get('copy_from'):
                raise web.HTTPError(400, "Cannot copy with PUT, only POST")
            if model.get('path') and (cm.is_hidden(path) or cm.is_hidden(model.get('path')))  and not cm.allow_hidden:
                raise web.HTTPError(400, f'Cannot create file or directory {path!r}')
            exists = yield maybe_future(self.contents_manager.file_exists(path))
            if exists:
                yield maybe_future(self._save(model, path))
            else:
                yield maybe_future(self._upload(model, path))
        else:
            yield maybe_future(self._new_untitled(path))

    @web.authenticated
    @gen.coroutine
    def delete(self, ws_uid, path=''):
        """delete a file in the given path"""
        cm = self.contents_manager

        if cm.is_hidden(path) and not cm.allow_hidden:
            raise web.HTTPError(400, f'Cannot delete file or directory {path!r}')

        self.log.warning('delete %s', path)
        yield maybe_future(cm.delete(path))
        self.set_status(204)
        self.finish()

class FilesHandler(IPythonHandler):
    @web.authenticated
    @gen.coroutine
    @interceptor
    def get(self, ws_uid, path, include_body=True) -> {"GRADE": ["BASIC", "PRO", "ENT"], "ROLE": ["ADMIN", "EDITOR","PORTALVIEWER"]}:
        # /files/ requests must originate from the same site
        self.check_xsrf_cookie()
        cm = FileContentsManager()
        if cm.is_hidden(path) and not cm.allow_hidden:
            self.log.info("Refusing to serve hidden file, via 404 Error")
            raise web.HTTPError(404)
        path = path.strip('/')
        if '/' in path:
            _, name = path.rsplit('/', 1)
        else:
            name = path
        model = yield maybe_future(cm.get(path, type='file', content=include_body))
        self.set_attachment_header(name)
        self.set_header('Content-Type', 'application/octet-stream')
        self.write(model['content'])
        self.flush()

class FilesEditHandler(IPythonHandler):

    def _finish_model(self, model, location=True):
        """Finish a JSON request with a model, setting relevant headers, etc."""
        if location:
            location = self.location_url(model['path'])
            self.set_header('Location', location)
        self.set_header('Content-Type', 'application/json')
        self.finish(json.dumps(model, default=json_default))

    @interceptor
    def get(self, ws_uid, path) -> {"GRADE": ["BASIC", "PRO", "ENT"], "ROLE": ["ADMIN", "EDITOR"]}:
        """
            :param excel_path: excel文件地址
            :param sheets_names: 列表传入指定sheet页 不传默认所有
            :return:
            """
        size = os.path.getsize(path)
        size = size/1024/1024
        if size > 10:
            ret_obj = {
                "code": 400,
                "msg": "文件超过限制（10M）！"
            }
            self._finish_model(ret_obj, location=False)
            return
        sheets_name = self.get_argument('sheets_name', '')
        s_list = path.split('.')
        extension = s_list[-1]
        #判断是否是csv 情况
        if extension == "csv":
            data = pd.read_csv(path)
        else:
            data = pd.read_excel(path)
            if sheets_name == "":
                data = pd.read_excel(path)
            else:
                data = pd.read_excel(path, sheets_name)
        temp_list = []
        type_list = []
        s = 1
        data = data.fillna("")
        #for column in data.iter_cols():
        for column in data.columns:
            type_dict = {}
            if column is not None and data[column].dtype is not None:
                type_dict['label'] = column
                type_dict['type'] = data[column].dtype.name
                type_dict['id'] = "cell_" + str(s)
            else:
                type_dict['label'] = "index"
                type_dict['type'] = data[column].dtype.name
                type_dict['id'] = "cell_" + str(s)
            s = s+1
            type_list.append(type_dict)
        '''for i in range(1, data.max_row):
            temp_dict = {}
            for j in range(1, data.max_column + 1):
                title_data = data.cell(row=1, column=j).value
                value_data = data.cell(row=i + 1, column=j).value
                if title_data is not None and value_data is not None:
                    temp_dict["cell_"+str(j)] = value_data
                elif title_data is None:
                    temp_dict["cell_" + str(j)] = value_data
                else:
                    continue
            temp_list.append(temp_dict)
        json_str = data.to_json(orient='records')
        #json_data = json.dumps(temp_list, indent=4, ensure_ascii=False)
        #json_data_type = json.dumps(type_list, indent=4, ensure_ascii=False)'''

        columns = [i["id"] for i in type_list]
        for i in data.itertuples():
            temp_list.append(dict(zip(columns, list(i[1:]))))
        date = {
            "data": temp_list,
            "colHeaders": type_list
        }
        ret_obj = {
            "code": 200,
            "data": date,
            "msg": "查询成功！"
        }
        self._finish_model(ret_obj, location=False)

class FilesSaveHandler(IPythonHandler):

    def _finish_model(self, model, location=True):
        """Finish a JSON request with a model, setting relevant headers, etc."""
        if location:
            location = self.location_url(model['path'])
            self.set_header('Location', location)
        self.set_header('Content-Type', 'application/json')
        self.finish(json.dumps(model, default=json_default))

    @interceptor
    def post(self, ws_uid, path) -> {"GRADE": ["BASIC", "PRO", "ENT"], "ROLE": ["ADMIN", "EDITOR"]}:
        try:
            params = json.loads(self.request.body)
            # 假设从前端传输过来的CSV格式文件字符串保存在变量 csv_string 中
            csv_string = params.get("csv_string")

            # 将CSV格式文件字符串转换为类似文件对象的 StringIO 对象
            csv_file = io.StringIO(csv_string)

            s_list = path.split('.')
            extension = s_list[-1]
            if extension == "csv":
                # 创建CSV文件读取器对象
                csv_reader = csv.reader(csv_file)
                # 打开要保存的本地文件，并将CSV数据写入其中
                with open(path, mode='w', newline='') as output_file:
                    csv_writer = csv.writer(output_file)
                    for row in csv_reader:
                        csv_writer.writerow(row)
            else:
                df = pd.read_csv(csv_file)
                df.to_excel(path, index=False)
            ret_obj = {
                "code": 200,
                "data": None,
                "msg": "保存成功！"
            }
            self._finish_model(ret_obj, location=False)
        except Exception as e:
            self.db_session.rollback()
            self.log.error(e)
            self.write_err("保存失败！")

class FilesFindHandler(IPythonHandler):

    def _finish_model(self, model, location=True):
        """Finish a JSON request with a model, setting relevant headers, etc."""
        if location:
            location = self.location_url(model['path'])
            self.set_header('Location', location)
        self.set_header('Content-Type', 'application/json')
        self.finish(json.dumps(model, default=json_default))

    @interceptor
    def get(self, ws_uid) -> {"GRADE": ["BASIC", "PRO", "ENT"], "ROLE": ["ADMIN", "EDITOR"]}:
        try:
            content = []
            file_name = self.get_argument('file_name', '')
            if file_name:
                for dirpath, dirnames, filenames in os.walk('/home'):
                    dirnames[:] = [d for d in dirnames if not d.startswith(".")]
                    for filename in filenames:
                        if file_name in filename:
                            obj = {}
                            #yield os.path.abspath(os.path.join(dirpath, filename))
                            # 使用 os.stat() 函数获取文件状态信息
                            stat_info = os.stat(f'{dirpath}/{filename}')
                            # 获取文件的更新时间（最后修改时间）
                            update_time = stat_info.st_mtime
                            # 将时间戳转换为日期时间格式
                            formatted_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(update_time))
                            size = stat_info.st_size
                            if size > 1024:
                                m = size / 1024
                                if m > 1024:
                                    size = str(Decimal(m / 1024).quantize(Decimal('0.00'))) + ' MB'
                                else:
                                    size = str(Decimal(m).quantize(Decimal('0.00'))) + ' KB'
                            else:
                                size = str(size) + ' bytes'
                            obj['name'] = filename
                            obj['file_path'] = f'{dirpath}/{filename}'
                            obj['update_time'] = formatted_time
                            obj['size'] = size
                            obj['type'] = "file"
                            content.append(obj)
                ret_obj = {
                    "code": 200,
                    "data": content,
                    "msg": "查询成功！"
                }
            else:
                ret_obj = {
                    "code": 200,
                    "data": [],
                    "msg": "查询成功！"
                }
            self._finish_model(ret_obj, location=False)
        except Exception as e:
            self.db_session.rollback()
            self.log.error(e)
            self.write_err("查询失败！")

_checkpoint_id_regex = r"(?P<checkpoint_id>[\w-]+)"
sheets_name = r"(?P<sheets_name>.*)"
_ws_uid = r"(?P<ws_uid>[^/]+)"
default_handlers = [
    (fr"/api/snb/node/filecontents/{_ws_uid}{path_regex}", ContentsHandler),
    (fr"/api/snb/node/filedownload/{_ws_uid}{path_regex}", FilesHandler),
    (fr"/api/snb/node/fileedit/{_ws_uid}{path_regex}", FilesEditHandler),
    (fr"/api/snb/node/filesave/{_ws_uid}{path_regex}", FilesSaveHandler),
    (fr"/api/snb/node/findfile/{_ws_uid}", FilesFindHandler),
]
