import re
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

from pathlib import Path

mypath = Path.cwd()


def row_Split(text, exp=None):
    # Split the text into lines
    lines = str(text).split('\n')
    abc = []  # Lines starting with numbers
    cab = []  # Lines not starting with numbers
    # Iterate through the lines and categorize them
    count = 1
    for line in lines:
        # Use a regular expression to check if the line starts with a number
        if re.match(r"^\d+\.", line.strip()):
            line = 'Step' + str(line)
            abc.append(line)
        else:
            if exp is None:
                if 'Expected Result' in line:
                    line = 'Step' + str(count) + '. ' + str(line)
                    count = count + 1
                cab.append(line)
    gabc = []
    for line in abc:
        gabc.append(line)
    return gabc, cab


def formulate_data(input_list):
    # Initialize variables for the final result and the current step
    result = []
    current_step = None
    # Iterate through the input list
    for item in input_list:
        if item.startswith('Step'):
            # If a step is found, append a new line and update the current step
            current_step = item
            result.append('\n' + current_step)
        elif item.strip() != '':
            # If the item is not empty after stripping, append it to the result
            result.append(item)
    # Join the result list into a single string
    rxt = '\n'.join(result)
    # Print or use the formatted_text as needed
    return rxt


def gmerge_ustry_name(sheet, data):
    # Merge the cells in the 'User_Story_Tcno' and 'Priority' columns
    def merge_cells_in_column(sheet, column):
        prev_value = None
        merge_start = 2
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=sheet[column][0].col_idx,
                                   max_col=sheet[column][0].col_idx):
            cell = row[0]
            cell_value = cell.value
            if cell_value != prev_value:
                if merge_start < row[0].row:
                    sheet.merge_cells(f'{column}{merge_start}:{column}{row[0].row - 1}')
                merge_start = row[0].row
                prev_value = cell_value
        if merge_start < sheet.max_row:
            sheet.merge_cells(f'{column}{merge_start}:{column}{sheet.max_row}')

    # Merge the cells in the 'User_Story_Tcno' and 'Priority' columns
    merge_cells_in_column(sheet, 'A')
    merge_cells_in_column(sheet, 'B')
    merge_cells_in_column(sheet, 'G')
    merge_cells_in_column(sheet, 'H')
    merge_cells_in_column(sheet, 'I')
    merge_cells_in_column(sheet, 'J')
    merge_cells_in_column(sheet, 'K')
    merge_cells_in_column(sheet, 'L')

    # Set alignment for merged cells
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def extract_excel_data_get(test_case_data_in, sheetname, test_case_data_out, gtColumns):
    global gabc
    tcid, tfeature, tcase, ttype, tprior, xsteps, texres = gtColumns
    gfeature = None
    input_file = test_case_data_in
    df = pd.read_excel(input_file)
    # Initialize an empty list to store DataFrames for each row
    dfs = []
    # Loop through the rows of the original DataFrame
    count = 1
    for index, row in df.iterrows():
        # Assuming the original DataFrame has 3 columns, you can access the data as follows:
        if tcid:
            col2 = row[tcid]
        else:
            idx = count + index
            col2 = 'Tcase_0' + str(idx)
        if tfeature and tcase is not None:
            if str(row[tfeature]) == 'nan':
                if gfeature:
                    col1 = str(gfeature) + ':' + str(row[tcase])
                else:
                    col1 = str(row[tcase])
            else:
                col1 = str(row[tfeature]) + ':' + str(row[tcase])
                gfeature = row[tfeature]
        else:
            col1 = str(row[tcase])
        if ttype:
            col9 = row[ttype]
        else:
            col9 = ttype
        if tprior:
            col10 = row[tprior]
        else:
            col10 = tprior
        if xsteps:
            getrows = row[xsteps]
            if texres:
                gabc, gcab = row_Split(getrows, exp='Yes')
                col8 = str(row[texres]) + '(' + str(index) + ')'
            else:
                gabc, gcab = row_Split(getrows)
                gtext = formulate_data(gcab)
                col8 = gtext
        else:
            col8 = index
        col4 = '***'
        col5 = '***'
        col6 = '***'
        col7 = index
        col11 = None
        col12 = None
        gtcase, ustry, tsteps, tdata, tseq, tactdesp, tresval, texperes, trun, tpri, tdepd, tinci = [], [], [], [], [], [], [], [], \
            [], [], [], []
        for xcol in range(len(gabc)):
            getcolxx_step = gabc[xcol]
            if getcolxx_step:
                col3 = getcolxx_step
                # Append the data to the new DataFrame
                new_row = pd.DataFrame(
                    {'Test_Case': [col1], 'User_Story_Tcno': [col2], 'Test_Steps': [col3], 'Test_data': [col4],
                     'Seq|Object_Repo_Ref|Action': [col5], 'Action Description': [col6],
                     'Results Validation[Automation]': [col7], 'Expected_Results[Manual]': [col8],
                     'RunType': [col9], 'Priority': [col10], 'Depend': [col11],
                     'IncidentIds': [col12]})

                # Append the new DataFrame to the list
                dfs.append(new_row)
                tsteps.append(col3)
        gtcase.append(col1)
        ustry.append(col2)
        tdata.append(col4)
        tseq.append(col5)
        tactdesp.append(col6)
        tresval.append(col7), texperes.append(col8), trun.append(col9), tpri.append(col10), tdepd.append(col11), \
            tinci.append(col12)
        tresval.clear()
    '''getcases = {'Test_Case': gtcase, 'User_Story_Tcno': ustry, 'Test_Steps': tsteps,
                'Seq|Object_Repo_Ref|Action': tseq, 'Action Description': tactdesp,
                'Results Validation[Automation]': tresval, 'Expected_Results[Manual]': texperes,
                'RunType': trun, 'Priority': tpri, 'Depend': tdepd, 'IncidentIds': tinci}'''
    # print(getcases)
    # Concatenate all DataFrames in the list
    new_df = pd.concat(dfs, ignore_index=True)
    output_file = test_case_data_out
    if os.path.isfile(output_file):
        wb = load_workbook(output_file)
        if sheetname in wb.sheetnames:
            ws = wb[sheetname]
            wb.remove(ws)
        ws = wb.create_sheet(title=sheetname)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheetname
    for row in dataframe_to_rows(new_df, index=False, header=True):
        ws.append(row)
    gmerge_ustry_name(sheet=ws, data=None)
    wb.save(output_file)
    print(f"Data from {input_file} has been split and saved to {output_file}.")


'''input_file = mypath / 'Testcase_raw.xlsx'
output_file = mypath / 'Testcase_raw_out.xlsx'
sheetnamex = 'FMS_Vehicle_Management'
gtColumns = 'ID', 'Section', 'Title', 'Type', 'Priority', 'Steps', None
extract_excel_data_get(test_case_data_in=input_file, sheetname=sheetnamex, test_case_data_out=output_file,
                       gtColumns=gtColumns)'''
