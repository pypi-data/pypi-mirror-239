import openpyxl
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.styles.borders import Side
from openpyxl.styles import Alignment


def create_excel_sheet_results(file_path, hTitle="OneFMS - Fleet Management System [ Portal ]",
                               sHtitle="Url: https://onefms-uat.zig.live/"):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the sheet name as "Object_repo"
    sheet.title = "Web_Results"

    # Headers for the Excel sheet
    headers = ['Tc_no', 'Features', 'UserStory', 'Test_Case', 'Test_Steps', 'Test_data',
               'Actual_Results', 'Expected_Results', 'Status', 'Screenshots', 'IncidentIds']

    # Set header values in the first row

    # Set formatting for the title row (first row)
    title_font = Font(name='Tahoma', size=20, bold=True, color="FFFFFF", underline="single")
    title_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_border = Border(top=Side(border_style="thin", color="000000"),
                          left=Side(border_style="thin", color="000000"),
                          right=Side(border_style="thin", color="000000"),
                          bottom=Side(border_style="thin", color="000000"))

    # Merge and center cells for the title row (first row)
    sheet.merge_cells('A1:K1')

    # Apply formatting for the title row (first row)
    title_cell = sheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    title_cell.border = title_border
    title_cell.value = hTitle

    # Set formatting for the subtitle row (second row)
    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    subtitle_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    subtitle_border = Border(top=Side(border_style="thin", color="000000"),
                             left=Side(border_style="thin", color="000000"),
                             right=Side(border_style="thin", color="000000"),
                             bottom=Side(border_style="thin", color="000000"))

    # Merge and center cells for the subtitle row (second row)
    sheet.merge_cells('E2:G2')

    # Apply formatting for the subtitle row (second row)
    subtitle_cell = sheet['E2']
    subtitle_cell.font = subtitle_font
    subtitle_cell.fill = subtitle_fill
    subtitle_cell.alignment = subtitle_alignment
    subtitle_cell.border = subtitle_border
    subtitle_cell.value = sHtitle

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    # Apply formatting for the subtitle row (second row)
    subtitle_cella = sheet['A2']
    subtitle_cella.font = subtitle_font
    subtitle_cella.fill = subtitle_fill
    subtitle_cella.alignment = subtitle_alignment
    subtitle_cella.border = subtitle_border
    subtitle_cella.value = 'Total Tests:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    subtitle_cellb = sheet['B2']
    subtitle_cellb.font = subtitle_font
    subtitle_cellb.fill = subtitle_fill
    subtitle_cellb.alignment = subtitle_alignment
    subtitle_cellb.border = subtitle_border
    subtitle_cellb.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    subtitle_cellc = sheet['C2']
    subtitle_cellc.font = subtitle_font
    subtitle_cellc.fill = subtitle_fill
    subtitle_cellc.alignment = subtitle_alignment
    subtitle_cellc.border = subtitle_border
    subtitle_cellc.value = 'Total Passed:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    subtitle_celld = sheet['D2']
    subtitle_celld.font = subtitle_font
    subtitle_celld.fill = subtitle_fill
    subtitle_celld.alignment = subtitle_alignment
    subtitle_celld.border = subtitle_border
    subtitle_celld.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid")
    subtitle_cellh = sheet['H2']
    subtitle_cellh.font = subtitle_font
    subtitle_cellh.fill = subtitle_fill
    subtitle_cellh.alignment = subtitle_alignment
    subtitle_cellh.border = subtitle_border
    subtitle_cellh.value = 'Total Failed:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid")
    subtitle_celli = sheet['I2']
    subtitle_celli.font = subtitle_font
    subtitle_celli.fill = subtitle_fill
    subtitle_celli.alignment = subtitle_alignment
    subtitle_celli.border = subtitle_border
    subtitle_celli.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="F4D25A", end_color="F4D25A", fill_type="solid")
    subtitle_cellj = sheet['J2']
    subtitle_cellj.font = subtitle_font
    subtitle_cellj.fill = subtitle_fill
    subtitle_cellj.alignment = subtitle_alignment
    subtitle_cellj.border = subtitle_border
    subtitle_cellj.value = 'Total Skipped:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="F4D25A", end_color="F4D25A", fill_type="solid")
    subtitle_cellk = sheet['k2']
    subtitle_cellk.font = subtitle_font
    subtitle_cellk.fill = subtitle_fill
    subtitle_cellk.alignment = subtitle_alignment
    subtitle_cellk.border = subtitle_border
    subtitle_cellk.value = '0'

    # Set formatting for the headers (third row)
    header_font = Font(bold=True, color="FFFFFF", underline="single")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

    sheet.append(headers)
    # Apply formatting for the headers (third row)
    for cell in sheet['A3:K3'][0]:
        cell.font = header_font
        cell.fill = header_fill

    # Apply borders to the first two rows except the header row
    row_border = Border(top=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

    for row in sheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = row_border

    # Save the Excel file
    workbook.save(file_path)


def create_excel_sheet_for_object_repo(file_path, app=None):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the sheet name as "Object_repo"
    sheet.title = "Object_Repo_BasePage"

    # Headers for the Excel sheet
    if app:
        headers = ['BasePage', 'Item', 'Locators', 'Elements1', 'Action', 'Description', 'FetchElements', 'Elements2']
    else:
        headers = ['BasePage', 'Item', 'Locators', 'Elements', 'Action', 'Description', 'FetchElements']

    # Set header values in the first row
    sheet.append(headers)

    # Apply formatting to the header row (underline, bold font, white text color, and red background)
    header_font = Font(bold=True, color="FFFFFF", underline="single")
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Apply borders to the first two rows except the header row
    row_border = Border(top=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

    for row in sheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = row_border

    # Save the Excel file
    workbook.save(file_path)


def create_excel_sheet_for_testcase(file_path):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the sheet name as "Object_repo"
    sheet.title = "FMS_Login"

    # Headers for the Excel sheet
    headers = ['Test_Case', 'User_Story_Tcno', 'Test_Steps', 'Test_data', 'Seq|Object_Repo_Ref|Action',
               'Action Description', 'Results Validation[Automation]', 'Expected_Results[Manual]', 'RunType', 'Case',
               'Depend', 'IncidentIds']

    # Set header values in the first row
    sheet.append(headers)

    # Apply formatting to the header row (underline, bold font, white text color, and red background)
    header_font = Font(bold=True, color="FFFFFF", underline="single")
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Apply borders to the first two rows except the header row
    row_border = Border(top=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

    for row in sheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = row_border

    # Save the Excel file
    workbook.save(file_path)


test_web_sample = """import os
import cdxg
import configparser
from cdxg import file_data
from wfs.utils.action_test_item import get_tags
from wfs.pages.object_repo_page import object_repox
from cdxg.logging import log
from wfs.utils.desired_actions import getdepend
from wfs.utils.common import Create_New_Report, get_results, get_line_dict

BASE_PATH = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
data_file_path = os.path.join(BASE_PATH, "conf.ini")
config = configparser.ConfigParser()
config.read(data_file_path)
test_case_data = config.get("Xldata", "test_case_data")
test_data_json = config.get("Xldata", "test_data_json")
test_object_repo = config.get('Xldata', 'test_object_repo')
test_tags = config.get("Tags", "run_type")
base_url = config.get("Url", "base_url")
xreport = config.get("Xldata", "test_report")
reportpath = Create_New_Report(report=xreport)


line = 7
end_line = None


class Test_0Web_FMS_Vehicle_Attribute(cdxg.TestCase):

	def start(self):
		self.lpage = object_repox(self.driver)
		self.baseUrl = f'{base_url}'

	@file_data(test_case_data, line=line, end_line=end_line, sheet="FMS_Vehicle_Attribute")
	def test_oneFMS_Vehicle_Attribute(self, testcase, ustory, teststeps, testdata, action, adescribe, aresults, exresults, rtype,
							case, depend, incidentids):
		""""""
		if rtype and case is not None:
			xtype = rtype + "," + case
		else:
			xtype = rtype if case is None else case
		if get_tags(xtype):
			log.info("TestCase : " + str(ustory) + "-" + str(testcase))
			log.info("Test_Tags : " + str(get_tags(xtype)))

		features = "FMS_Vehicle_Attribute"
		self.maxDiff = None
		if "skip" not in xtype and xtype == get_tags(xtype) and xtype is not None:
			dependdata = testcase, ustory, teststeps, testdata, action, adescribe, aresults, exresults, case, depend
			gtres = getdepend(test_case_data, reportpath, features, test_object_repo, test_data_json, dependdata, self.baseUrl, self.driver, ptname=None)
			for xgres in gtres:
				if 'Error' in xgres or gtres == 'Y' or xgres == 'Y':
					get_results(reportpath, features, ustory, testcase, teststeps, testdata, xgres, 
										exresults, gtres, lpage=self.lpage, results='FAILED', fontx='FC2C03', sshots=features, incidentids=incidentids)
					self.xFail(gtres)
					break
				if xgres.endswith('YES'):
					if '>>' in str(aresults) and not str(aresults).startswith('>>') and not str(aresults).endswith('>>'):
						aresults = get_line_dict(ptname=None, dname=aresults)
					acresults = str(xgres)+'>>'+str(aresults)
					get_results(reportpath, features, ustory, testcase, teststeps, testdata, acresults, 
										exresults, gtres, lpage=self.lpage, results='PASSED', fontx='35FC03', sshots=None, incidentids=incidentids)
					break
		else:
			if rtype == "skip":
				self.xSkip(testcase + ": Testcase Skipped, Due to not much information")
				get_results(reportpath, features, ustory, testcase, teststeps, testdata, 'Testcase Skipped', 
												exresults, lpage=self.lpage, results='SKIPPED', fontx='F4D25A', sshots=None, incidentids=incidentids)
			else:
				self.skipTest(reason="Test execution based on Tags :" + str(test_tags) + ": Excludes the rest")
"""

test_zigapp_sample = """import os
import cdxg
import configparser
from cdxg import file_data
from wfs.utils.action_test_item import get_tags
from cdxg.logging import log
from wfs.utils.common import Create_New_Report, get_results, getdepend

BASE_PATH = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
data_file_path = os.path.join(BASE_PATH, "conf.ini")
config = configparser.ConfigParser()
config.read(data_file_path)
test_case_data = config.get("MobileHost", "test_case_data")
test_data_json = config.get("MobileHost", "test_data_json")
test_object_repo = config.get('MobileHost', 'test_object_repo')
test_tags = config.get("Tags", "run_type")
xreport = config.get("MobileHost", "test_report")
reportpath = Create_New_Report(report=xreport)


line = 2
end_line = None


class Test_Mobile_APP_Login(cdxg.TestCase):

	@file_data(test_case_data, line=line, end_line=end_line, sheet="APP_Login")
	def test_zigAPP_Login(self, testcase, ustory, teststeps, testdata, action, adescribe, aresults, exresults, rtype,
							case, depend, incidentids):
		""""""
		if get_tags(rtype):
			log.info("Test_Tags: " + str(get_tags(rtype)))
		features = "APP_Login"
		self.maxDiff = None

		if rtype != "skip" and rtype == get_tags(rtype) and rtype is not None:
			dependdata = testcase, ustory, teststeps, testdata, action, adescribe, aresults, exresults, case, depend
			gtres = getdepend(test_case_data, reportpath, features, test_object_repo, test_data_json, dependdata, url=None, driver=None, ptname="Android")
			if 'Error' in gtres or gtres == 'Y':
				get_results(reportpath, features, ustory, testcase, teststeps, testdata, str(gtres), 
										exresults, results='FAILED', fontx='FC2C03', sshots=features, incidentids=incidentids)
				self.xFail(gtres)
			if gtres == 'YES':
				get_results(reportpath, features, ustory, testcase, teststeps, testdata, aresults, 
										exresults, results='PASSED', fontx='35FC03', sshots=None, incidentids=incidentids)
		else:
			get_results(reportpath, features, ustory, testcase, teststeps, testdata, 'Testcase Skipped', 
												exresults, results='SKIPPED', fontx='F4D25A', sshots=None, incidentids=incidentids)
			if rtype == "skip":
				self.xSkip(testcase + ": Testcase Skipped, Due to not much information")
			else:
				self.skipTest(reason="Test execution based on Tags :" + str(test_tags) + ": Excludes the rest")
"""

run_test = """import os
import platform
from cdxg.logging import log
import cdxg
import subprocess
import time
import argparse
from configparser import ConfigParser
from selenium.webdriver import ChromeOptions
from wfs.utils.action_test_item import generate_test_object_repo
from wfs.utils.common import create_excel_testcase
from pathlib import Path

cObject = ConfigParser()
cObject.read('conf.ini')

Mypath = Path.cwd()

if __name__ == '__main__':
    # run test file
    parser = argparse.ArgumentParser(description='Add Arguments')
    parser.add_argument("--t", help="Add run type Tags")
    parser.add_argument("--b", help="Add Browser/App")
    parser.add_argument("--s", help="Give Sheetname")
    parser.add_argument("--l", help="Give Executelines")
    args = parser.parse_args()
    tags_add = cObject['Tags']
    tags_add['run_type'] = args.t
    with open('conf.ini', 'w') as cf:
        cObject.write(cf)

    sheetname = args.s
    exeLine = args.l

    if args.b == 'gc':
        subprocess.Popen(['mitmdump', '-s', 'proxy.py'])  # , '--web-port', '8081'])
        log.info('MitmDump and Proxy Connection Started......')
        time.sleep(1)
        proxy_address = "127.0.0.1:8080"
        chrome_options = ChromeOptions()
        chrome_options.add_argument(f'--proxy-server={proxy_address}')
        chrome_options.add_argument('--ignore-certificate-errors')
        chrome_options.add_experimental_option("excludeSwitches", ['enable-automation', 'enable-logging'])
        # chrome_options.add_experimental_option("mobileEmulation", {"deviceName": "iPad Pro"})
        browser = {
            "browser": args.b,
            "options": chrome_options,
        }
    else:
        download_dir = Mypath / 'reports' / 'download'
        if args.b == 'chrome':
            chrome_options = ChromeOptions()
            chrome_options.add_argument('--enable-logging')
            chrome_options.set_capability('goog:loggingPrefs', {'performance': 'ALL'})
            chrome_options.add_experimental_option("prefs", {
                "download.default_directory": str(download_dir),
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": True
            })
            browser = {
                "browser": args.b,
                "options": chrome_options,
            }
        else:
            browser = args.b

    if args.b in ['Android', 'iOS']:
        # file_pathxx = ["./test_data/mobile/get_text_item.txt", "./test_data/mobile/text_data_repo.txt"]
        file_pathxx = Mypath / 'test_data' / 'mobile' / 'get_text_item.txt'
        for fpath in file_pathxx:
            with open(fpath, 'r+') as file:
                file.truncate(0)
        case_data = cObject.get('MobileHost', 'test_case_data')
        test_case_data = Mypath / 'test_data' / 'mobile' / case_data
        saved_case_data = Mypath / 'test_data' / case_data
        test_pfix = cObject.get('MobileHost', 'test_prefix')
        generate_test_object_repo(test_case_data, pfix=test_pfix, dx='Mobile', ptname=args.b, sheetname=sheetname,
                                  exeLine=exeLine)
        description = 'App : ' + args.b + ' / RunType : ' + cObject.get('Tags', 'run_type')
        pyfile = Mypath / 'test_dir' / 'test_zigapp.py'
        cdxg.main(str(pyfile), tester='Shafee Syed', description=description, 
                  title='Fleet Management Services[oneFMS]', report='result.html')
        file_path = test_case_data
        try:
            os.remove(file_path)
            print(f"{file_path} has been successfully deleted.")
        except OSError as e:
            print(f"Error deleting {file_path}: {e}")
    else:
        fpath = Mypath / 'test_data' / 'web' / 'get_text_item.txt'
        with open(fpath, 'r+') as file:
            file.truncate(0)
        case_data = cObject.get('Xldata', 'test_case_data')
        test_case_data = Mypath / 'test_data' / 'web' / case_data
        saved_case_data = Mypath / 'test_data' / case_data
        test_pfix = cObject.get('Xldata', 'test_prefix')
        create_excel_testcase(test_case_data=saved_case_data, saved_case_data=test_case_data)
        generate_test_object_repo(test_case_data, pfix=test_pfix, sheetname=sheetname, exeLine=exeLine)
        description = 'Browser : ' + args.b + ' / RunType : ' + cObject.get('Tags', 'run_type')
        cdxg.TestCase.maxDiff = None
        pyfile = Mypath / 'test_dir' / 'test_onefms.py'
        cdxg.main(str(pyfile), browser=browser, tester='Shafee Syed', description=description, 
                  title='Fleet Management Services[oneFMS]', report='result.html')
        file_path = test_case_data
        try:
            os.remove(file_path)
            print(f"{file_path} has been successfully deleted.")
        except OSError as e:
            print(f"Error deleting {file_path}: {e}")
    if args.b == 'gc':
        system = platform.system()
        if system == 'Windows':
            subprocess.Popen(['taskkill', '/F', '/IM', 'mitmdump'])
        else:
            subprocess.Popen(['kill', '-f', 'mitmdump'])
        log.info('MitmDump and proxy Connection Closed......')
"""

test_wproxy = """from mitmproxy import ctx
from mitmproxy.addonmanager import AddonManager
from mitmproxy import http
from mitmproxy.tools.main import mitmproxy
from pathlib import Path
import os, json
import platform

mypath = Path.cwd()


def getspliter(apix, method):
    spltapi = apix.split('/v1.0/')[1]
    if '/' in spltapi and '?' not in spltapi:
        spltapi = str(spltapi).replace('/', '_')
    if '?' in spltapi:
        spltapi = str(spltapi).split('?')[0]
    return spltapi + '_' + method.lower()


class APIEndpointLogger:
    def __init__(self):
        self.api_endpoints = set()

    def request(self, flow: http.HTTPFlow) -> None:
        try:
            if flow.request.path.startswith("/fms/api/v1.0/"):
                print("API Request:" + str(flow.request.path))
                # Save the request content to a file
                fnamex = getspliter(apix=flow.request.path, method=flow.request.method)
                request_filename = mypath / 'test_data' / 'web' / 'pxy' / 'request'
                os.makedirs(request_filename, exist_ok=True)
                request_filename = str(request_filename / (fnamex + '.json'))
                getrequest = json.loads(flow.request.content.decode())
                with open(request_filename, 'w') as file:
                    # file.write(json.dump(getrequest))
                    json.dump(getrequest, file, ensure_ascii=False, indent=4)
        except json.JSONDecodeError as e:
            return

    def response(self, flow: http.HTTPFlow) -> None:
        try:
            if flow.request.path.startswith("/fms/api/v1.0/"):
                self.api_endpoints.add((flow.request.path, flow.response.content))
                fnamex = getspliter(apix=flow.request.path, method=flow.request.method)
                response_filename = mypath / 'test_data' / 'web' / 'pxy' / 'response'
                os.makedirs(response_filename, exist_ok=True)
                response_filename = str(response_filename / (fnamex + '.json'))
                getresponse = json.loads(flow.response.content.decode())
                with open(response_filename, 'w') as file:
                    # file.write(json.dump(getresponse))
                    json.dump(getresponse, file, ensure_ascii=False, indent=4)
        except json.JSONDecodeError as e:
            return


def done(self) -> None:
    print("API Endpoints:")
    for endpoint in self.api_endpoints:
        print(f"Path: {endpoint[0]}")
        print(f"Content: {endpoint[1]}")
        print("-" * 20)


addons = [
    APIEndpointLogger()
]


def start_proxy():
    mitmproxy(["-s", __file__])


if __name__ == "__main__":
    start_proxy()
"""

test_mproxy = """from mitmproxy import ctx
from mitmproxy.addonmanager import AddonManager
from mitmproxy import http
from mitmproxy.tools.main import mitmproxy
from pathlib import Path
import os, json
import platform

mypath = Path.cwd()


# https://interapps-uat.cdgtaxi.com.sg
def getspliter(apix, method):
    spltapi = apix.split('/dcp-cms/rest/')[1]
    if '/' in spltapi and '?' not in spltapi:
        spltapi = str(spltapi).replace('/', '_')
    if '?' in spltapi:
        spltapi = str(spltapi).split('?')[0]
    return spltapi + '_' + method.lower()


class APIEndpointLogger:
    def __init__(self):
        self.api_endpoints = set()

    def request(self, flow: http.HTTPFlow) -> None:
        try:
            if flow.request.path.startswith("/dcp-cms/rest/"):
                print("API Request:" + str(flow.request.path))
                # Save the request content to a file
                fnamex = getspliter(apix=flow.request.path, method=flow.request.method)
                request_filename = mypath / 'test_data' / 'mobile' / 'pxy' / 'request'
                os.makedirs(request_filename, exist_ok=True)
                request_filename = str(request_filename / (fnamex + '.json'))
                getrequest = json.loads(flow.request.content.decode())
                with open(request_filename, 'w') as file:
                    # file.write(json.dump(getrequest))
                    json.dump(getrequest, file, ensure_ascii=False, indent=4)
        except json.JSONDecodeError as e:
            return

    def response(self, flow: http.HTTPFlow) -> None:
        try:
            if flow.request.path.startswith("/dcp-cms/rest/"):
                self.api_endpoints.add((flow.request.path, flow.response.content))
                fnamex = getspliter(apix=flow.request.path, method=flow.request.method)
                response_filename = mypath / 'test_data' / 'mobile' / 'pxy' / 'response'
                os.makedirs(response_filename, exist_ok=True)
                response_filename = str(response_filename / (fnamex + '.json'))
                getresponse = json.loads(flow.response.content.decode())
                with open(response_filename, 'w') as file:
                    # file.write(json.dump(getresponse))
                    json.dump(getresponse, file, ensure_ascii=False, indent=4)
        except json.JSONDecodeError as e:
            return


def done(self) -> None:
    print("API Endpoints:")
    for endpoint in self.api_endpoints:
        print(f"Path: {endpoint[0]}")
        print(f"Content: {endpoint[1]}")
        print("-" * 20)


addons = [
    APIEndpointLogger()
]


def start_proxy():
    mitmproxy(["-s", __file__])


if __name__ == "__main__":
    start_proxy()
"""

conf_requires = """[Url]
base_url = https://onefms-uat.zig.live/

[Xldata]
test_case_data = fms_test_case_and_data.xlsx
test_object_repo = fms_object_repo.xlsx
test_obj_sheet1 = Object_Repo_BasePage
test_obj_col1 = BasePage
test_prefix = FMS_
test_report = FMS_Report
api_target_list = /api/v1.0/
xheaders = 

[Tags]
run_type = Low

[MobileHost]
test_case_data = app_test_case_and_data.xlsx
test_object_repo = app_object_repo.xlsx
test_prefix = APP_
test_report = APP_Report
local_host = http://localhost:4723/wd/hub
android_current_device_pax = emulator-5554
android_current_device_drv = emulator-5556
android_app_path_pax = ShopFront.apk
android_app_path_driver = Driverapp.apk
android_activity_pax = com.codigo.comfort
android_activity_drv = com.codigo.cdgdriver.uat
android_platform_version = 10
ios_app_path_pax = ShopFront.ipa
ios_app_path_driver = Driverapp.ipa
ios_activity = 
ios_current_device = iphone 13
ios_platform_version = 15.6
ios_activity_pax = com.codigo.comfort
ios_activity_drv = com.codigo.cdgdriver.uat
"""

test_requires = """asyncio
ffmpeg-python
"""

addon_uipage = """# This  uipage.py file used for create additional functions and methods for webelements to interact with UI pages, 
# If in which doesnt covered in default UI webelements 
# Use in a way as same as Class, constructor and function defined..

from cdxg.logging import log
from selenium.common import StaleElementReferenceException
from wfs.utils.assertions import Assertion
from selenium.webdriver.common.keys import Keys
import traceback, time, re


class Page_reference(object):

    def __init__(self, lpage, elementget, testitem, aresults):
        self.lpage = lpage
        self.elementget = elementget
        self.testitem = testitem
        self.aresults = aresults

    def addon_type(self):
        try:
            self.lpage.locators_ready(greadyx=self.elementget).clear()
            self.lpage.locators_ready(greadyx=self.elementget).send_keys(self.testitem)
            log.info('Enter data as : ' + str(self.testitem))
            return 'Enter data as : ' + str(self.testitem)
        except Exception as e:
            return 'Error ' + str(e)"""


addon_appage = """# This  appage.py file used for create additional functions and methods for mobileelements to interact with Mobile app pages, 
# If in which doesnt covered in default  mobile elements
# Use in a way as same as Class, constructor and function defined.. 

from cdxg.logging import log
from selenium.common import StaleElementReferenceException
from wfs.utils.assertions import Assertion
from selenium.webdriver.common.keys import Keys
import traceback, time, re


class Page_reference(object):

    def __init__(self, lpage, elementget, testitem, aresults):
        self.lpage = lpage
        self.elementget = elementget
        self.testitem = testitem
        self.aresults = aresults

    def addon_type(self):
        try:
            pass
        except Exception as e:
            return 'Error ' + str(e)"""