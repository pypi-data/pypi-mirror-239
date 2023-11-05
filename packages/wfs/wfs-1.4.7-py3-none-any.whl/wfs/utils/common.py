import shutil
import pyautogui
import openpyxl
import time
import random
import string
import re
from pathlib import Path
from wfs.utils.parseexcel import ParseExcel
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.styles.borders import Side
from openpyxl.styles import Alignment
from configparser import ConfigParser

mypath = Path.cwd()  # .parent
cObject = ConfigParser()
cObject.read('conf.ini')


def id_generator(size=6, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


def get_results(xlcreate, features, userstory, testcase, teststeps, testdata, actResults, expResults, aresults, results,
                fontx, sshots, incidentids, lpage, comments='{Step1: Go to Url, Step2: Yes}'):
    global screens
    reportpath = ParseExcel(excel_path=xlcreate)
    reportpath.set_sheet_by_name('Web_Results')
    get_total_rows = reportpath.get_max_row_no()
    reportpath.write_cell_content_colored(2, 2, '=COUNTA(A4:A10000)')
    reportpath.write_cell_content_colored(2, 4, '=COUNTIF(I4:I10000, "✅")')
    reportpath.write_cell_content_colored(2, 9, '=COUNTIF(I4:I10000, "❌")')
    reportpath.write_cell_content_colored(2, 11, '=COUNTIF(I4:I10000, "⚠️")')
    aresults = '\n'.join(aresults)
    try:
        reportpath.write_cell_content_colored(get_total_rows + 1, 1, get_total_rows - 2)
        reportpath.write_cell_content_colored(get_total_rows + 1, 2, features)
        reportpath.write_cell_content_colored(get_total_rows + 1, 3, userstory)
        reportpath.write_cell_content_colored(get_total_rows + 1, 4, testcase)
        reportpath.write_cell_content_colored(get_total_rows + 1, 5, teststeps)
        reportpath.write_cell_content_colored(get_total_rows + 1, 6, testdata)
        reportpath.write_cell_content_colored(get_total_rows + 1, 7, str(aresults))
        reportpath.getcomment(get_total_rows + 1, 7, str(actResults))
        reportpath.write_cell_content_colored(get_total_rows + 1, 8, str(expResults))
        if results == 'PASSED':
            reportpath.write_cell_content_colored(get_total_rows + 1, 9, '✅', font=fontx)
        elif results == 'SKIPPED':
            reportpath.write_cell_content_colored(get_total_rows + 1, 9, '⚠️', font=fontx)
        else:
            reportpath.write_cell_content_colored(get_total_rows + 1, 9, '❌', font=fontx)
        # reportpath.write_cell_content_colored(get_total_rows + 1, 9, results, font=fontx)
        if results == 'FAILED':
            screens = lpage.screen_shots(screenshot_path=str(sshots + '_' + str(get_total_rows + 1)))
            reportpath.sHyperlink(get_total_rows + 1, 10, sshot=sshots + '_' + str(get_total_rows + 1),
                                  sshotpath=str(screens))
        else:
            reportpath.write_cell_content_colored(get_total_rows + 1, 10, sshots)
        reportpath.write_cell_content_colored(get_total_rows + 1, 11, incidentids)
    except Exception as e:
        if ValueError:
            for xna in actResults:
                reportpath.write_cell_content_colored(get_total_rows + 1, 7, xna)
        print(str(e))


def Create_New_Report(report):
    otime = time.strftime("%Y-%m-%d_%H%M%S", time.localtime())
    reportfolderpath = report + "_" + otime + ".xlsx"
    # shutil.copy(mypath / 'reports' / 'web_results.xlsx', mypath / 'reports' / reportfolderpath)
    reportPath = mypath / 'reports' / reportfolderpath
    create_excel_sheet_results(file_path=reportPath)
    return mypath / 'reports' / reportfolderpath


def screenshots(sname):
    screenshot = pyautogui.screenshot()
    snamex = mypath / 'reports' / 'screenshots' / sname
    screenshot.save(snamex)
    return snamex


def get_set_data_attribute(gtstring):
    global stg1, stg2
    input_string = gtstring
    # Define a regular expression pattern to match the desired values
    pattern = r'^(.*?)\[(.*?)\]$'
    # Use re.match to find the pattern in the input string
    match = re.match(pattern, input_string)

    if match:
        # Extract the desired values
        stg1 = match.group(1)
        stg2 = match.group(2)
    return stg1, stg2


def get_line_dict(ptname, dname):
    ptn = 'web'
    if ptname:
        ptn = 'mobile'
    with open(mypath / 'test_data' / ptn / 'get_text_item.txt', 'r') as frd:
        lines = frd.readlines()
    data = lines
    # print(data)

    result = {}

    for item in data:
        parts = item.strip().split('|')
        if len(parts) == 5:
            if parts[2] != '***':
                result.setdefault(parts[0], {})[parts[2]] = parts[1]
    # print(result)
    result = {k: {key: value for key, value in v.items() if value != '***'} for k, v in result.items() if v}

    if all(char in dname for char in (',', '>>')):
        yall = []
        getxall = dname.split(',')
        for xall in getxall:
            if '>>' in xall:
                dxt, nmxt = xall.split('>>')
                yall.append(result[dxt][nmxt])
            else:
                yall.append(xall)
    else:
        dxt, nmxt = dname.split('>>')
        yall = result[dxt][nmxt]
    return yall


def get_expected(keyitem):
    getlen = []
    if '*' in keyitem:
        getsplit = str(keyitem).split('*')
        for xlen in getsplit:
            getlen.append(xlen)
    else:
        getlen = keyitem  # .append(keyitem)
    return getlen


def get_check_list(gtx, artx):
    itm = []
    mitm = []
    gtx_items = gtx
    if type(gtx) is not list:
        gtx_items = gtx.split(',')
    no_data_available = True
    alldata = 'Y'
    for item in gtx_items:
        if item not in artx:
            alldata = 'N'
            mitm.append(item)
            no_data_available = False
    if no_data_available:
        alldata = alldata
        mitm = 'Y'
    itm.append(alldata)
    return itm, gtx_items, mitm


def getcheckorder(order_list, expected_order):
    result = None
    if type(expected_order) is not list:
        if ',' in expected_order and '<<' not in expected_order:
            expected_order = expected_order.split(',')

            def check_order(order_list, expected_order):
                pn = []
                exp = 0
                for xorder in order_list:
                    try:
                        if xorder == expected_order[exp]:
                            exp = exp
                            pn.append('Y')
                        elif xorder != expected_order[exp]:
                            exp += 1
                            if xorder == expected_order[exp]:
                                exp = exp
                                pn.append('Y')
                            else:
                                pn.append('N')
                        else:
                            exp += 1
                            pn.append('N')
                    except Exception:
                        pn.append('N')
                for yorder in expected_order:
                    if yorder not in order_list:
                        pn.append('N')
                return list(set(pn)), expected_order, None

            result = check_order(order_list, expected_order)
        else:
            if str(expected_order).startswith('<<'):
                expected_order = expected_order.split('<<')[1]
            else:
                expected_order = expected_order.split('<<')[0]
            result = get_check_list(gtx=expected_order, artx=order_list)
    else:
        result = get_check_list(gtx=expected_order, artx=order_list)
    return result


def extract_details(lpage, gAitem, gready, aresults, gTestitem, ptname):
    from wfs.pages.base_page import get_ordering_table
    gcount = None
    if str(aresults).startswith('>>'):
        aresults = get_ordering_table(lpage, gAitem, gready, aresults, gTestitem)
    else:
        if '>>' in gTestitem and not str(gTestitem).startswith('>>') and not str(gTestitem).endswith('>>'):
            gTestitem = get_line_dict(ptname, dname=gTestitem)
        if '>>' in str(aresults) and not str(aresults).startswith('>>') and not str(aresults).endswith('>>'):
            aresults = get_line_dict(ptname, dname=aresults)
        if str(aresults).endswith('>>'):
            aresults, gntx = str(aresults).split('>>')
            gcount = 1
        if str(gTestitem).endswith('>>'):
            gTestitem, gntx = str(gTestitem).split('>>')
            gcount = 1
        if '*' in aresults:
            aresults = str(aresults).split('*')
    return gTestitem, aresults, ptname, gcount


def get_test_case_extraction(new_workbook, data, sheet_name=None):
    global new_sheet
    output_data = [('Test_Case', 'User_Story_Tcno', 'Test_Steps', 'Test_data', 'Seq|Object_Repo_Ref|Action',
                    'Action Description', 'Results Validation[Automation]', 'Expected_Results[Manual]', 'RunType',
                    'Priority', 'Depend', 'IncidentIds')]

    current_test_case = None
    current_user_story_no = None
    current_test_steps = ""
    current_test_data = ""
    current_test_seq = ""
    current_test_actdesp = ""
    current_actual_results = None
    current_expected_results = None
    current_runtype = None
    current_priority = None
    current_depend = None
    current_incidents = None
    priorityx = None
    runtypex = None

    for row in data[1:]:
        test_case_id, user_story_no, test_steps, test_data, test_seq, test_actdesp, actual_results, expected_results, \
            runtype, priority, depend, incidents = row
        if test_case_id:
            if current_test_case:
                output_data.append((current_test_case, current_user_story_no, current_test_steps, current_test_data,
                                    current_test_seq, current_test_actdesp, current_actual_results,
                                    current_expected_results, current_runtype, current_priority, current_depend,
                                    current_incidents))
            current_test_case = test_case_id
            current_user_story_no = user_story_no
            current_test_steps = test_steps
            current_test_data = test_data
            current_test_seq = test_seq
            current_test_actdesp = test_actdesp
            current_actual_results = actual_results
            current_expected_results = expected_results
            current_runtype = runtype
            current_priority = priority
            current_depend = depend
            current_incidents = incidents
            if current_runtype is not None:
                runtypex = current_runtype
            else:
                current_runtype = runtypex
            if current_priority is not None:
                priorityx = current_priority
            else:
                current_priority = priorityx
        else:
            current_test_steps += "\n" + str(test_steps)
            current_test_data += "\n" + str(test_data)
            current_test_seq += "\n" + str(test_seq)
            current_test_actdesp += "\n" + str(test_actdesp)
            if actual_results:
                current_actual_results = actual_results
            if expected_results:
                current_expected_results = expected_results
            if runtype:
                current_runtype = runtype
            if priority:
                current_priority = priority
            if depend:
                current_depend = depend
            if incidents:
                current_incidents = incidents
            if current_runtype is not None:
                runtypex = current_runtype
            else:
                current_runtype = runtypex
            if current_priority is not None:
                priorityx = current_priority
            else:
                current_priority = priorityx

    # Add the last test case
    if current_test_case:
        output_data.append((current_test_case, current_user_story_no, current_test_steps, current_test_data,
                            current_test_seq, current_test_actdesp, current_actual_results,
                            current_expected_results, current_runtype, current_priority, current_depend,
                            current_incidents))

    # Print the transformed data
    hy = 0
    for row in output_data:
        if row == ('Test_Case', 'User_Story_Tcno', 'Test_Steps', 'Test_data', 'Seq|Object_Repo_Ref|Action',
                   'Action Description', 'Results Validation[Automation]', 'Expected_Results[Manual]', 'RunType',
                   'Priority', 'Depend', 'IncidentIds'):
            new_sheet = new_workbook.create_sheet(title=sheet_name[hy])
            hy += 1
        new_sheet.append(row)


def create_excel_testcase(test_case_data, saved_case_data):
    # Load the original Excel workbook
    original_workbook = openpyxl.load_workbook(test_case_data, data_only=True)
    # Create a new Excel workbook
    new_workbook = openpyxl.Workbook()
    # Iterate through sheets in the original workbook
    xk = []
    sht = []
    for sheet_name in original_workbook.sheetnames:
        original_sheet = original_workbook[sheet_name]
        for row in original_sheet.iter_rows(values_only=True):
            xk.append(row)
        sht.append(sheet_name)
    get_test_case_extraction(new_workbook, data=xk, sheet_name=sht)
    # Remove the default sheet created in the new workbook
    new_workbook.remove(new_workbook.active)
    # Save the new workbook to a new Excel file
    new_workbook.save(saved_case_data)
    # Close both workbooks
    original_workbook.close()
    new_workbook.close()
    time.sleep(1)


def getlocator_direct(action, desptn):
    if action[0] in ['xpath', 'id_', 'id', 'css']:
        elementIdentity, locatorIdentity, actionIdentity, felements, itemIdentity = action
        actionlist = locatorIdentity, elementIdentity, actionIdentity, felements, itemIdentity
        tuple_list = [tuple(actionlist)]
        new_tuple = tuple_list[0][:3] + (desptn,) + tuple_list[0][3:]
        tuple_list[0] = new_tuple
        return tuple_list
    else:
        return


def check_word_order(sentence):
    gword = []
    xwords = ['find', 'click', 'enter', 'then', 'and', 'select', 'go']
    words = sentence.split()
    if 'then' in gword:
        result = []
        sublist = []
        for item in gword:
            if item != 'then':
                sublist.append(item)
            else:
                result.append(sublist)
            sublist = []
        result.append(sublist)
        gword = result
        return gword
    else:
        for x in words:
            for y in xwords:
                if x == y: gword.append(x)
        return gword


def getaction_datalist(action, steps_line, testdata, actiondesc):
    getaction = action.split('\n')
    testdata = testdata.split('\n')
    actiondesc = actiondesc.split('\n')
    action = getaction[steps_line]
    return getaction[steps_line], testdata[steps_line], actiondesc[steps_line]


def get_action_dict(action):
    xmt = []
    if len(action) >= 2:
        for xaction in action:
            xmt.append(xaction)
        action = ''.join(xmt)
    else:
        action = action[0]

    actionlist = {'find': 'text', 'click': 'click', 'enter': 'type', 'Find': 'text', 'Click': 'click', 'Enter': 'type',
                  'findandenter': 'type', 'clickandenter': 'type', 'findandclick': 'click', 'findandselect': 'click',
                  'select': 'click',
                  'selecttext': 'link_text', 'selectoption': 'link_text', 'go': 'url'}
    elementslist = {'type': 'S', 'click': 'S', 'text': 'M', 'select': 'S', 'link_text': 'M', 'url': None}
    return actionlist[action], action, elementslist[actionlist[action]]


def check_word_order_list(sentence):
    gword = []
    xwords = ['find', 'click', 'enter', 'then', 'and', 'select']
    words = sentence.split()
    for x in words:
        for y in xwords:
            if x == y:
                gword.append(x)
    if 'click' in words and 'enter' in words and 'find' not in words:
        if words.index('click') < words.index('enter'):
            return "'click' appears before 'enter' after in the sentence: '{}'".format(sentence), gword
        return "'enter' appear before 'click' after in the sentence: '{}'".format(sentence), gword
    elif 'find' in words and 'enter' in words and 'click' not in words:
        if words.index('find') < words.index('enter'):
            return "'find' appears before 'enter' after in the sentence: '{}'".format(sentence), gword
        return "'enter' appear before 'find' after in the sentence: '{}'".format(sentence), gword
    elif 'find' in words and 'click' in words and 'enter' not in words:
        if words.index('find') < words.index('click'):
            return "'find' appears before 'click' in the sentence: '{}'".format(sentence), gword
        return "'click' appear before 'find' after in the sentence: '{}'".format(sentence), gword
    elif 'find' in words and 'click' not in words and 'enter' not in words:
        return "'only find appears: '{}'".format(sentence), gword
    elif 'click' in words and 'enter' not in words and 'find' not in words:
        return "'only find appears: '{}'".format(sentence), gword
    elif 'enter' in words and 'find' not in words and 'click' not in words:
        return "'only find appears: '{}'".format(sentence), gword
    else:
        return "'Others...........: '{}'".format(sentence), gword


def create_excel_sheet_results(file_path, hTitle="OneFMS - Fleet Management System [ Portal ]",
                               sHtitle="Url: https://onefms-uat.zig.live/"):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the sheet name as "Object_repo"
    sheet.title = "Web_Results"

    # Headers for the Excel sheet
    headers = ['Tc_no', 'Features', 'UserStory', 'Test_Case', 'Test_Steps', 'Test_data',
               'Actual_Results', 'Expected_Results', 'Status', 'Screenshots', 'IncidentIds']

    # Set header values in the first row

    # Set formatting for the title row (first row)
    title_font = Font(name='Tahoma', size=20, bold=True, color="FFFFFF", underline="single")
    title_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_border = Border(top=Side(border_style="thin", color="000000"),
                          left=Side(border_style="thin", color="000000"),
                          right=Side(border_style="thin", color="000000"),
                          bottom=Side(border_style="thin", color="000000"))

    # Merge and center cells for the title row (first row)
    sheet.merge_cells('A1:K1')

    # Apply formatting for the title row (first row)
    title_cell = sheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    title_cell.border = title_border
    title_cell.value = hTitle

    # Set formatting for the subtitle row (second row)
    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    subtitle_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    subtitle_border = Border(top=Side(border_style="thin", color="000000"),
                             left=Side(border_style="thin", color="000000"),
                             right=Side(border_style="thin", color="000000"),
                             bottom=Side(border_style="thin", color="000000"))

    # Merge and center cells for the subtitle row (second row)
    sheet.merge_cells('E2:G2')

    # Apply formatting for the subtitle row (second row)
    subtitle_cell = sheet['E2']
    subtitle_cell.font = subtitle_font
    subtitle_cell.fill = subtitle_fill
    subtitle_cell.alignment = subtitle_alignment
    subtitle_cell.border = subtitle_border
    subtitle_cell.value = sHtitle

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    # Apply formatting for the subtitle row (second row)
    subtitle_cella = sheet['A2']
    subtitle_cella.font = subtitle_font
    subtitle_cella.fill = subtitle_fill
    subtitle_cella.alignment = subtitle_alignment
    subtitle_cella.border = subtitle_border
    subtitle_cella.value = 'Total Tests:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    subtitle_cellb = sheet['B2']
    subtitle_cellb.font = subtitle_font
    subtitle_cellb.fill = subtitle_fill
    subtitle_cellb.alignment = subtitle_alignment
    subtitle_cellb.border = subtitle_border
    subtitle_cellb.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    subtitle_cellc = sheet['C2']
    subtitle_cellc.font = subtitle_font
    subtitle_cellc.fill = subtitle_fill
    subtitle_cellc.alignment = subtitle_alignment
    subtitle_cellc.border = subtitle_border
    subtitle_cellc.value = 'Total Passed:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    subtitle_celld = sheet['D2']
    subtitle_celld.font = subtitle_font
    subtitle_celld.fill = subtitle_fill
    subtitle_celld.alignment = subtitle_alignment
    subtitle_celld.border = subtitle_border
    subtitle_celld.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid")
    subtitle_cellh = sheet['H2']
    subtitle_cellh.font = subtitle_font
    subtitle_cellh.fill = subtitle_fill
    subtitle_cellh.alignment = subtitle_alignment
    subtitle_cellh.border = subtitle_border
    subtitle_cellh.value = 'Total Failed:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid")
    subtitle_celli = sheet['I2']
    subtitle_celli.font = subtitle_font
    subtitle_celli.fill = subtitle_fill
    subtitle_celli.alignment = subtitle_alignment
    subtitle_celli.border = subtitle_border
    subtitle_celli.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="F4D25A", end_color="F4D25A", fill_type="solid")
    subtitle_cellj = sheet['J2']
    subtitle_cellj.font = subtitle_font
    subtitle_cellj.fill = subtitle_fill
    subtitle_cellj.alignment = subtitle_alignment
    subtitle_cellj.border = subtitle_border
    subtitle_cellj.value = 'Total Skipped:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="F4D25A", end_color="F4D25A", fill_type="solid")
    subtitle_cellk = sheet['k2']
    subtitle_cellk.font = subtitle_font
    subtitle_cellk.fill = subtitle_fill
    subtitle_cellk.alignment = subtitle_alignment
    subtitle_cellk.border = subtitle_border
    subtitle_cellk.value = '0'

    # Set formatting for the headers (third row)
    header_font = Font(bold=True, color="FFFFFF", underline="single")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

    sheet.append(headers)
    # Apply formatting for the headers (third row)
    for cell in sheet['A3:K3'][0]:
        cell.font = header_font
        cell.fill = header_fill

    # Apply borders to the first two rows except the header row
    row_border = Border(top=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

    for row in sheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = row_border

    # Save the Excel file
    workbook.save(file_path)
